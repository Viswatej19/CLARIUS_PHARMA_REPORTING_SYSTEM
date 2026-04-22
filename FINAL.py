import streamlit as st
import pandas as pd
from datetime import datetime
import io
import re

# ── Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

# ── PDF
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

st.set_page_config(layout="wide", page_title="Clarius Pharma - Invoice")

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

SELLER = """M/s CLARIUS PHARMA.
D.NO.19-1-358/4A. LECTURERS COLONY.
SRI RAMASAI NILAYAM PEDDAPURAM - 533437.
DL NO'S : 20B-AP/04/06/2017-137922.
21B-AP/04/06/2017-137923
GSTIN - 37AKYPB8654K1ZB"""

BUYERS = {
    "Mahalaxmi":          "TO, MAHALAXMI MEDICALS.\nRAMAMAHAL LADIES GATE ROAD. R.R. PET. ELURU-2.\nGSTIN NO- 37AJDPK5246N1ZT.\nD L NO- 339/AP/WG/E2011/R",
    "Balaji":             "TO, M/s BALAJI M/G/S. TADEPALLIGUDEM\nGSTIN NO-37BOPPP4153R1ZX.\n20&21-AP/05/2015-125403;125404",
    "Ramaswamy":          "TO: RAMASWAMY MEDICALS\nC/O DR. POTUMUDI SRINIWAS\nNEAR SUBBAMMADEVI SCHOOL, ELURU-2\nPH: 8008143357\nGSTIN: 37ASHPM3995K1ZZ\nDL NO: 140457-AP/05/01/2017 / 140458-AP/05/01/2017",
    "Laxmi Medicals":     "TO M/s LAXMI MEDICALS RETAIL SHOP,\nRAMARAOPET, KAKINADA.\nGSTIN-37ABIPV1833D1ZM.",
    "Manikanta":          "TO M/S SRI MANIKANTA MEDICALS.\nKOTHA ROAD, ELURU.\nD L NO: 54 54\nGSTIN: 37ACPPD6515N1Z5",
    "Madhura":            "TO, MADHURA MEDICALS YELESWARAM.\nGSTIN NO-37ABEFM6531H1Z7.\n20&21-985/AP/EG/R/2010/R",
    "Datta Sai Medicals": "TO, SRI DATTA SAI MEDICAL STORES.\nR.R.PETA, ELURU.\nDL.NO: 140/AP/WG/E/2009/R\nGSTIN: 37BRAPP1812M1ZO",
}

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

HDR_COLOR  = "D0F5FC"   # user-requested header colour
ALT_COLOR  = "EEF9FB"   # very light tint for alt rows
SUM_COLOR  = "F0FAFE"   # summary section bg
BORDER_CLR = "8BCDD8"

def thin_border(left=True, right=True, top=True, bottom=True):
    s = Side(style="thin", color="8BCDD8")
    n = Side(style=None)
    return Border(
        left=s if left else n,
        right=s if right else n,
        top=s if top else n,
        bottom=s if bottom else n,
    )

def hdr_fill():
    return PatternFill("solid", fgColor=HDR_COLOR)

def alt_fill():
    return PatternFill("solid", fgColor=ALT_COLOR)

def sum_fill():
    return PatternFill("solid", fgColor=SUM_COLOR)

def style_cell(cell, value=None, bold=False, size=9, align="center",
               valign="center", fill=None, wrap=False, color="000000",
               border=True, num_format=None):
    if value is not None:
        cell.value = value
    cell.font = Font(name="Calibri", size=size, bold=bold, color=color)
    cell.alignment = Alignment(
        horizontal=align, vertical=valign,
        wrap_text=wrap
    )
    if fill:
        cell.fill = fill
    if border:
        cell.border = thin_border()
    if num_format:
        cell.number_format = num_format

def merge_style(ws, cell_range, value=None, bold=False, size=9,
                align="center", valign="center", fill=None,
                wrap=True, color="000000", border=True, num_format=None):
    ws.merge_cells(cell_range)
    c1 = ws[cell_range.split(":")[0]]
    style_cell(c1, value=value, bold=bold, size=size, align=align,
               valign=valign, fill=fill, wrap=wrap, color=color,
               border=border, num_format=num_format)
    # Apply border to all cells in the merged range
    if border:
        from openpyxl.utils import rows_from_range
        for row in rows_from_range(cell_range):
            for coord in row:
                ws[coord].border = thin_border()

def apply_outer_border(ws, min_row, min_col, max_row, max_col):
    """Apply a clean outer border around a rectangular range."""
    thick = Side(style="medium", color="4AAFBC")
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            left   = thick if c == min_col   else cell.border.left
            right  = thick if c == max_col   else cell.border.right
            top    = thick if r == min_row   else cell.border.top
            bottom = thick if r == max_row   else cell.border.bottom
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

def num_fmt(v, dec=2):
    if v == 0:
        return 0
    return round(v, dec)

def create_excel(df, buyer_key, buyer_text, inv_no, date_val, transport, lr_no, lr_date):
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left   = 0.4
    ws.page_margins.right  = 0.4
    ws.page_margins.top    = 0.5
    ws.page_margins.bottom = 0.5

    # ── Column widths (A..Q = 17 cols)
    col_widths = {
        'A': 5,   'B': 22,  'C': 9,   'D': 9,   'E': 12,
        'F': 8,   'G': 7,   'H': 7,   'I': 9,   'J': 8,
        'K': 9,   'L': 11,  'M': 6,   'N': 10,  'O': 6,
        'P': 10,  'Q': 12,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ══ ROW 1: Title ══════════════════════════════════════════════════════════
    ws.row_dimensions[1].height = 22
    merge_style(ws, "A1:Q1", value="GST INVOICE",
                bold=True, size=14, align="center", valign="center",
                fill=PatternFill("solid", fgColor="C5F0F8"), border=True)

    # ══ ROW 2: Seller | Buyer | Info ══════════════════════════════════════════
    ws.row_dimensions[2].height = 85
    merge_style(ws, "A2:E2", value=SELLER,
                bold=False, size=8, align="left", valign="top",
                fill=PatternFill("solid", fgColor="F0FAFE"), border=True)

    merge_style(ws, "F2:M2", value=buyer_text,
                bold=True, size=8.5, align="left", valign="top",
                fill=PatternFill("solid", fgColor="F0FAFE"), border=True)

    info_text = (f"Inv No : {inv_no}\n"
                 f"Date   : {date_val}\n"
                 f"Transport : {transport}\n"
                 f"Lr No  : {lr_no}\n"
                 f"Lr Date: {lr_date}")
    merge_style(ws, "N2:Q2", value=info_text,
                bold=False, size=8, align="left", valign="top",
                fill=PatternFill("solid", fgColor="F0FAFE"), border=True)

    # ══ ROW 3-4: Column Headers ════════════════════════════════════════════════
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 14

    hf = hdr_fill()
    hdr_style = dict(bold=True, size=8, align="center", valign="center", fill=hf, border=True)

    # Merged headers spanning rows 3-4
    for rng, label in [
        ("A3:A4",   "#"),
        ("B3:B4",   "PRODUCT NAME"),
        ("C3:C4",   "PACKING"),
        ("D3:D4",   "HSN CODE"),
        ("E3:E4",   "BATCH NO"),
        ("F3:F4",   "EXP"),
        ("G3:G4",   "QTY"),
        ("H3:H4",   "FREE"),
        ("I3:I4",   "PTR"),
        ("J3:J4",   "DISC"),
        ("K3:K4",   "MRP"),
        ("L3:L4",   "TAXABLE AMT"),
    ]:
        merge_style(ws, rng, value=label, **hdr_style)

    # CGST spans M3:N3, then sub-headers in row 4
    merge_style(ws, "M3:N3", value="CGST", **hdr_style)
    style_cell(ws["M4"], value="%",     **hdr_style)
    style_cell(ws["N4"], value="T.AMT", **hdr_style)

    # SGST spans O3:P3
    merge_style(ws, "O3:P3", value="SGST", **hdr_style)
    style_cell(ws["O4"], value="%",     **hdr_style)
    style_cell(ws["P4"], value="T.AMT", **hdr_style)

    # TOTAL AMT spans Q3:Q4
    merge_style(ws, "Q3:Q4", value="TOTAL AMT", **hdr_style)

    # ══ DATA ROWS ══════════════════════════════════════════════════════════════
    data_start = 5
    n_rows = len(df)

    for i, row in df.iterrows():
        r = data_start + i
        ws.row_dimensions[r].height = 16
        bg = alt_fill() if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        dc = dict(size=9, border=True, fill=bg)

        style_cell(ws.cell(r, 1),  value=i+1,                align="center", **dc)
        style_cell(ws.cell(r, 2),  value=row["Product"],     align="left",   **dc)
        style_cell(ws.cell(r, 3),  value=row["Packing"],     align="center", **dc)
        style_cell(ws.cell(r, 4),  value=row["HSN"],         align="center", **dc)
        style_cell(ws.cell(r, 5),  value=row["Batch"],       align="center", **dc)
        style_cell(ws.cell(r, 6),  value=row["EXP"],         align="center", **dc)
        style_cell(ws.cell(r, 7),  value=num_fmt(row["Qty"],0),  align="center", num_format="#,##0",    **dc)
        style_cell(ws.cell(r, 8),  value=num_fmt(row["Free"],0), align="center", num_format="#,##0",    **dc)
        style_cell(ws.cell(r, 9),  value=num_fmt(row["PTR"]),    align="center", num_format="#,##0.00", **dc)
        style_cell(ws.cell(r, 10), value=num_fmt(row["Discount"]),align="center",num_format="#,##0.00", **dc)
        style_cell(ws.cell(r, 11), value=num_fmt(row["MRP"]),    align="center", num_format="#,##0.00", **dc)
        style_cell(ws.cell(r, 12), value=num_fmt(row["Taxable"]),align="center", num_format="#,##0.00", **dc)
        style_cell(ws.cell(r, 13), value=float(row["CGST%"]), align="center", num_format="0.0", **dc)
        style_cell(ws.cell(r, 14), value=num_fmt(row["CGST Amt"]),align="center",num_format="#,##0.00", **dc)
        style_cell(ws.cell(r, 15), value=float(row["SGST%"]), align="center", num_format="0.0", **dc)
        style_cell(ws.cell(r, 16), value=num_fmt(row["SGST Amt"]),align="center",num_format="#,##0.00", **dc)
        style_cell(ws.cell(r, 17), value=num_fmt(row["Total"]),  align="center", num_format="#,##0.00", **dc)

    # ══ SUMMARY SECTION ════════════════════════════════════════════════════════
    sum_row = data_start + n_rows  # row immediately after products

    # Blank separator row
    ws.row_dimensions[sum_row].height = 6
    for c in range(1, 18):
        ws.cell(sum_row, c).fill = PatternFill("solid", fgColor="D0F5FC")

    s1 = sum_row + 1   # first summary row

    # Calculate summaries
    subtotal    = df["Taxable"].sum()
    cgst_tot    = df["CGST Amt"].sum()
    sgst_tot    = df["SGST Amt"].sum()
    gst_tot     = cgst_tot + sgst_tot
    total_disc  = (df["Qty"] * df["Discount"]).sum()
    net         = subtotal + gst_tot
    tot_items   = len(df[df["Taxable"] > 0])
    tot_units   = df["Qty"].sum()

    gst_rates   = [0, 5, 12, 18, 28]
    gst_summary = {}
    for _, row in df.iterrows():
        rate = int(row["CGST%"] + row["SGST%"])
        gst_summary.setdefault(rate, {"tax": 0, "cgst": 0, "sgst": 0})
        gst_summary[rate]["tax"]  += row["Taxable"]
        gst_summary[rate]["cgst"] += row["CGST Amt"]
        gst_summary[rate]["sgst"] += row["SGST Amt"]

    # Header row for summary
    for r_off, height in enumerate([14, 14, 14, 14, 14, 14, 14]):
        ws.row_dimensions[s1 + r_off].height = height

    sf = sum_fill()

    # GST breakdown header
    shdr = dict(bold=True, size=8, align="center", fill=hdr_fill(), border=True)
    merge_style(ws, f"A{s1}:B{s1}", value="CGST% VALUE",  **shdr)
    merge_style(ws, f"C{s1}:D{s1}", value="CGST AMT",     **shdr)
    merge_style(ws, f"E{s1}:F{s1}", value="SGST VALUE",   **shdr)
    merge_style(ws, f"G{s1}:H{s1}", value="SGST AMT",     **shdr)
    merge_style(ws, f"I{s1}:J{s1}", value="CASES",        **shdr)
    merge_style(ws, f"K{s1}:L{s1}", value="",             **shdr)
    merge_style(ws, f"M{s1}:N{s1}", value="AMOUNT",       **shdr)
    merge_style(ws, f"O{s1}:Q{s1}", value="VALUE",        **shdr)

    # 5 GST rate rows
    for idx, rate in enumerate(gst_rates):
        r = s1 + 1 + idx
        d = gst_summary.get(rate, {"tax": 0, "cgst": 0, "sgst": 0})
        bg = alt_fill() if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        sc = dict(size=8.5, fill=bg, border=True)

        merge_style(ws, f"A{r}:B{r}", value=f"{rate} % :", align="left",   **sc)
        merge_style(ws, f"C{r}:D{r}", value=num_fmt(d["tax"]),  align="right", num_format="#,##0.00", **sc)
        merge_style(ws, f"E{r}:F{r}", value=num_fmt(d["cgst"]), align="right", num_format="#,##0.00", **sc)
        merge_style(ws, f"G{r}:H{r}", value=num_fmt(d["tax"]),  align="right", num_format="#,##0.00", **sc)
        merge_style(ws, f"I{r}:J{r}", value=num_fmt(d["sgst"]), align="right", num_format="#,##0.00", **sc)

        # Right side labels & values
        if idx == 0:
            merge_style(ws, f"K{r}:L{r}", value="NO OF ITEMS :", align="left",  fill=sf, bold=False, size=8.5, border=True)
            merge_style(ws, f"M{r}:N{r}", value=tot_items,        align="right", fill=sf, bold=True,  size=8.5, border=True)
            merge_style(ws, f"O{r}:Q{r}", value="SUB TOTAL :",    align="left",  fill=sf, bold=False, size=8.5, border=True)
            # Will fill value below
        elif idx == 1:
            merge_style(ws, f"K{r}:L{r}", value="NO OF UNITS :", align="left",  fill=sf, bold=False, size=8.5, border=True)
            merge_style(ws, f"M{r}:N{r}", value=num_fmt(tot_units,0), align="right", fill=sf, bold=True, size=8.5, border=True, num_format="#,##0")
            merge_style(ws, f"O{r}:Q{r}", value="LESS DISC :",   align="left",  fill=sf, bold=False, size=8.5, border=True)
        elif idx == 2:
            merge_style(ws, f"K{r}:L{r}", value="",              fill=sf, size=8.5, border=True)
            merge_style(ws, f"M{r}:N{r}", value="",              fill=sf, size=8.5, border=True)
            merge_style(ws, f"O{r}:Q{r}", value="GST AMT :",     align="left",  fill=sf, bold=False, size=8.5, border=True)
        elif idx == 3:
            merge_style(ws, f"K{r}:L{r}", value="",              fill=sf, size=8.5, border=True)
            merge_style(ws, f"M{r}:N{r}", value="",              fill=sf, size=8.5, border=True)
            merge_style(ws, f"O{r}:Q{r}", value="CR AMT :",      align="left",  fill=sf, bold=False, size=8.5, border=True)
        elif idx == 4:
            merge_style(ws, f"K{r}:L{r}", value="",              fill=sf, size=8.5, border=True)
            merge_style(ws, f"M{r}:N{r}", value="",              fill=sf, size=8.5, border=True)
            merge_style(ws, f"O{r}:Q{r}", value="",              fill=sf, size=8.5, border=True)

    # Fill right-side values for SUB TOTAL etc. (column Q of the right block)
    # We placed the label in O col; we need a value col — let's add values in a separate pass
    # The right block is O:Q — label + value. Let's redo as label in O:P and value in Q
    val_rows = {
        s1+1: ("SUB TOTAL",  num_fmt(subtotal),   "#,##0.00"),
        s1+2: ("LESS DISC",  num_fmt(total_disc),  "#,##0.00"),
        s1+3: ("GST AMT",    num_fmt(gst_tot),     "#,##0.00"),
        s1+4: ("CR AMT",     0,                    "#,##0.00"),
        s1+5: ("",           "",                   None),
    }
    for r, (label, val, fmt) in val_rows.items():
        # Clear and redo the O:Q range properly
        ws.unmerge_cells(f"O{r}:Q{r}")
        merge_style(ws, f"O{r}:P{r}", value=label, align="left", fill=sf, bold=False, size=8.5, border=True)
        style_cell(ws[f"Q{r}"], value=val, align="right", fill=sf, bold=True,
                   size=8.5, border=True, num_format=fmt or "General")

    # NET PAYABLE row
    net_row = s1 + 6
    ws.row_dimensions[net_row].height = 16
    net_fill = PatternFill("solid", fgColor="B8EEF7")
    merge_style(ws, f"A{net_row}:N{net_row}", value="",
                fill=net_fill, bold=True, size=9, border=True)
    merge_style(ws, f"O{net_row}:P{net_row}", value="NET PAYABLE",
                fill=net_fill, bold=True, size=10, align="right", border=True)
    style_cell(ws[f"Q{net_row}"], value=num_fmt(round(net)),
               bold=True, size=10, align="right",
               fill=net_fill, border=True, num_format="#,##0.00")

    # Print area
    last_row = net_row
    ws.print_area = f"A1:Q{last_row}"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# PDF GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

PDF_HDR_BG = colors.HexColor("#D0F5FC")
PDF_HDR_FG = colors.HexColor("#003344")
PDF_ALT    = colors.HexColor("#EEF9FB")
PDF_NRM    = colors.white
PDF_SUM_BG = colors.HexColor("#F0FAFE")
PDF_BORDER = colors.HexColor("#4AAFBC")
PDF_NET_BG = colors.HexColor("#B8EEF7")

def make_styles():
    ps = lambda name, **kw: ParagraphStyle(name, **kw)
    return {
        "seller": ps("seller", fontName="Helvetica-Bold", fontSize=7.5, leading=10, spaceAfter=0),
        "buyer":  ps("buyer",  fontName="Helvetica-Bold", fontSize=7.5, leading=10, spaceAfter=0),
        "info":   ps("info",   fontName="Helvetica", fontSize=7.5, leading=10, spaceAfter=0, alignment=TA_RIGHT),
        "hdr":    ps("hdr",    fontName="Helvetica-Bold", fontSize=6.5, leading=8, spaceAfter=0, alignment=TA_CENTER),
        "cell":   ps("cell",   fontName="Helvetica", fontSize=7, leading=9, spaceAfter=0, alignment=TA_CENTER),
        "cellL":  ps("cellL",  fontName="Helvetica", fontSize=7, leading=9, spaceAfter=0),
        "title":  ps("title",  fontName="Helvetica-Bold", fontSize=13, leading=16, spaceAfter=2, alignment=TA_CENTER),
        "sumL":   ps("sumL",   fontName="Helvetica", fontSize=7.5, leading=10, spaceAfter=0),
        "sumR":   ps("sumR",   fontName="Helvetica", fontSize=7.5, leading=10, spaceAfter=0, alignment=TA_RIGHT),
        "sumB":   ps("sumB",   fontName="Helvetica-Bold", fontSize=7.5, leading=10, spaceAfter=0, alignment=TA_RIGHT),
    }

def pnum(v, dec=2):
    if v == 0:
        return "-"
    return f"{v:,.{dec}f}"

def create_pdf(df, buyer_text, inv_no, date_val, transport, lr_no, lr_date):
    buf = io.BytesIO()
    M = 12 * mm
    PW, PH = landscape(A4)

    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=M, rightMargin=M,
        topMargin=10*mm, bottomMargin=10*mm,
    )
    W = PW - 2*M
    ST = make_styles()
    story = []

    # Title
    story.append(Paragraph("GST INVOICE", ST["title"]))
    story.append(HRFlowable(width="100%", thickness=2, color=PDF_HDR_BG, spaceAfter=4))

    # Header block
    seller_para = Paragraph(SELLER.replace("\n", "<br/>"), ST["seller"])
    buyer_para  = Paragraph(buyer_text.replace("\n", "<br/>"), ST["buyer"])
    info_text   = (f"<b>Inv No  :</b> {inv_no}<br/>"
                   f"<b>Date    :</b> {date_val}<br/>"
                   f"<b>Transport:</b> {transport}<br/>"
                   f"<b>Lr No   :</b> {lr_no}<br/>"
                   f"<b>Lr Date :</b> {lr_date}")
    info_para = Paragraph(info_text, ST["info"])

    hdr_tbl = Table([[seller_para, buyer_para, info_para]],
                    colWidths=[W*0.27, W*0.45, W*0.28])
    hdr_tbl.setStyle(TableStyle([
        ("BOX",          (0,0), (-1,-1), 0.8, PDF_BORDER),
        ("INNERGRID",    (0,0), (-1,-1), 0.5, PDF_BORDER),
        ("VALIGN",       (0,0), (-1,-1), "TOP"),
        ("BACKGROUND",   (0,0), (-1,-1), PDF_SUM_BG),
        ("LEFTPADDING",  (0,0), (-1,-1), 5),
        ("RIGHTPADDING", (0,0), (-1,-1), 5),
        ("TOPPADDING",   (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0), (-1,-1), 5),
    ]))
    story.append(hdr_tbl)
    story.append(Spacer(1, 5))

    # Products table — only actual product rows, no blank padding
    H  = lambda t: Paragraph(t, ST["hdr"])
    C  = lambda t: Paragraph(str(t), ST["cell"])
    CL = lambda t: Paragraph(str(t), ST["cellL"])

    headers = [
        H("#"), H("PRODUCT NAME"), H("PACKING"), H("HSN"),
        H("BATCH"), H("EXP"), H("QTY"), H("FREE"), H("PTR"),
        H("DISC"), H("MRP"), H("TAXABLE"), H("CGST\n%"),
        H("CGST\nAMT"), H("SGST\n%"), H("SGST\nAMT"), H("TOTAL"),
    ]
    cw_raw   = [4, 18, 7, 7, 8, 5, 5, 5, 7, 6, 7, 8, 5, 6, 5, 6, 8]
    col_widths = [W * u / sum(cw_raw) for u in cw_raw]

    rows = [headers]
    for i, row in df.iterrows():
        rows.append([
            C(i+1), CL(row["Product"]), C(row["Packing"]), C(row["HSN"]),
            C(row["Batch"]), C(row["EXP"]),
            C(pnum(row["Qty"], 0)), C(pnum(row["Free"], 0)),
            C(pnum(row["PTR"])), C(pnum(row["Discount"])),
            C(pnum(row["MRP"])), C(pnum(row["Taxable"])),
            C(pnum(row["CGST%"], 1)), C(pnum(row["CGST Amt"])),
            C(pnum(row["SGST%"], 1)), C(pnum(row["SGST Amt"])),
            C(pnum(row["Total"])),
        ])

    n = len(rows)
    prod_tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    prod_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  PDF_HDR_BG),
        ("TEXTCOLOR",     (0,0), (-1,0),  PDF_HDR_FG),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0),  6.5),
        ("ALIGN",         (0,0), (-1,0),  "CENTER"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        *[("BACKGROUND",  (0,r), (-1,r), PDF_ALT if r % 2 == 0 else PDF_NRM)
          for r in range(1, n)],
        ("BOX",           (0,0), (-1,-1), 0.8, PDF_BORDER),
        ("INNERGRID",     (0,0), (-1,-1), 0.3, PDF_BORDER),
        ("TOPPADDING",    (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING",   (0,0), (-1,-1), 2),
        ("RIGHTPADDING",  (0,0), (-1,-1), 2),
    ]))
    story.append(prod_tbl)
    story.append(Spacer(1, 6))

    # Summary section
    subtotal   = df["Taxable"].sum()
    cgst_tot   = df["CGST Amt"].sum()
    sgst_tot   = df["SGST Amt"].sum()
    gst_tot    = cgst_tot + sgst_tot
    total_disc = (df["Qty"] * df["Discount"]).sum()
    net        = subtotal + gst_tot
    tot_items  = len(df[df["Taxable"] > 0])
    tot_units  = df["Qty"].sum()

    gst_summary = {}
    for _, row in df.iterrows():
        rate = int(row["CGST%"] + row["SGST%"])
        gst_summary.setdefault(rate, {"tax": 0, "cgst": 0, "sgst": 0})
        gst_summary[rate]["tax"]  += row["Taxable"]
        gst_summary[rate]["cgst"] += row["CGST Amt"]
        gst_summary[rate]["sgst"] += row["SGST Amt"]

    SL = lambda t: Paragraph(t, ST["sumL"])
    SR = lambda t: Paragraph(t, ST["sumR"])
    SB = lambda t: Paragraph(f"<b>{t}</b>", ST["sumB"])

    # GST breakdown table
    gst_data = [[
        Paragraph("<b>GST%</b>", ST["hdr"]),
        Paragraph("<b>TAX VALUE</b>", ST["hdr"]),
        Paragraph("<b>CGST AMT</b>", ST["hdr"]),
        Paragraph("<b>SGST VALUE</b>", ST["hdr"]),
        Paragraph("<b>SGST AMT</b>", ST["hdr"]),
    ]]
    for rate in [0, 5, 12, 18, 28]:
        d = gst_summary.get(rate, {"tax": 0, "cgst": 0, "sgst": 0})
        gst_data.append([SL(f"{rate}%"), SR(pnum(d["tax"])), SR(pnum(d["cgst"])),
                         SR(pnum(d["tax"])), SR(pnum(d["sgst"]))])

    gst_tbl = Table(gst_data, colWidths=[W*0.06, W*0.1, W*0.09, W*0.1, W*0.09])
    gst_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), PDF_HDR_BG),
        ("TEXTCOLOR",     (0,0), (-1,0), PDF_HDR_FG),
        ("BOX",           (0,0), (-1,-1), 0.8, PDF_BORDER),
        ("INNERGRID",     (0,0), (-1,-1), 0.3, PDF_BORDER),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING",   (0,0), (-1,-1), 3),
        ("RIGHTPADDING",  (0,0), (-1,-1), 3),
        *[("BACKGROUND",  (0,r), (-1,r), PDF_ALT if r % 2 == 0 else PDF_NRM)
          for r in range(1, 6)],
    ]))

    # Middle: counts
    mid_tbl = Table([
        [SL("CASES:"),         SR("")],
        [SL("No. of Items:"),  SB(str(tot_items))],
        [SL("No. of Units:"),  SB(pnum(tot_units, 0))],
        [SL("Due Date:"),      SR("")],
        [SL("Note:"),          SR("")],
    ], colWidths=[W*0.1, W*0.09])
    mid_tbl.setStyle(TableStyle([
        ("BOX",           (0,0), (-1,-1), 0.8, PDF_BORDER),
        ("INNERGRID",     (0,0), (-1,-1), 0.3, PDF_BORDER),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("BACKGROUND",    (0,0), (-1,-1), PDF_SUM_BG),
        ("TOPPADDING",    (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING",   (0,0), (-1,-1), 4),
        ("RIGHTPADDING",  (0,0), (-1,-1), 4),
    ]))

    # Right: financials
    right_tbl = Table([
        [SL("SUB TOTAL"),  SB(f"Rs. {pnum(subtotal)}")],
        [SL("LESS DISC"),  SR(f"Rs. {pnum(total_disc)}")],
        [SL("GST AMT"),    SR(f"Rs. {pnum(gst_tot)}")],
        [SL("CR AMT"),     SR("Rs. 0.00")],
        [SL(""),           SR("")],
        [Paragraph("<b>NET PAYABLE</b>", ST["sumB"]),
         Paragraph(f"<b>Rs. {pnum(round(net))}</b>", ST["sumB"])],
    ], colWidths=[W*0.12, W*0.14])
    right_tbl.setStyle(TableStyle([
        ("BOX",           (0,0), (-1,-1), 0.8, PDF_BORDER),
        ("INNERGRID",     (0,0), (-1,-1), 0.3, PDF_BORDER),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("BACKGROUND",    (0,0), (-1,-1), PDF_SUM_BG),
        ("BACKGROUND",    (0,5), (-1,5), PDF_NET_BG),
        ("TOPPADDING",    (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING",   (0,0), (-1,-1), 4),
        ("RIGHTPADDING",  (0,0), (-1,-1), 4),
    ]))

    sp = W * 0.02
    summary_outer = Table(
        [[gst_tbl, "", mid_tbl, "", right_tbl]],
        colWidths=[W*0.44, sp, W*0.19, sp, W*0.26],
    )
    summary_outer.setStyle(TableStyle([
        ("VALIGN",        (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING",   (0,0), (-1,-1), 0),
        ("RIGHTPADDING",  (0,0), (-1,-1), 0),
        ("TOPPADDING",    (0,0), (-1,-1), 0),
        ("BOTTOMPADDING", (0,0), (-1,-1), 0),
    ]))
    story.append(summary_outer)

    # Footer: signature only
    story.append(Spacer(1, 6))
    story.append(HRFlowable(width="100%", thickness=0.8, color=PDF_BORDER))
    footer_tbl = Table(
        [[Paragraph(""), Paragraph("For <b>CLARIUS PHARMA</b><br/>Authorised Signatory", ST["info"])]],
        colWidths=[W*0.7, W*0.3],
    )
    footer_tbl.setStyle(TableStyle([
        ("VALIGN",      (0,0), (-1,-1), "TOP"),
        ("TOPPADDING",  (0,0), (-1,-1), 3),
        ("LEFTPADDING", (0,0), (-1,-1), 0),
    ]))
    story.append(footer_tbl)

    doc.build(story)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<style>
    .block-container { padding-top: 1rem; }
    div[data-testid="stTextInput"] input { font-size: 13px; }
</style>
""", unsafe_allow_html=True)

st.markdown("## 🧾 Clarius Pharma — Invoice Generator")

col1, col2, col3 = st.columns([2, 2, 1])
with col1:
    st.text_area("FROM", SELLER, height=140)
with col2:
    buyer_key = st.selectbox("Select Buyer", list(BUYERS.keys()))
    st.text_area("TO", BUYERS[buyer_key], height=140)
with col3:
    today     = datetime.today().strftime("%d.%m.%Y")
    inv_no    = st.text_input("Inv No")
    name      = st.text_input("File Name")
    date_val  = st.text_input("Date", value=today)
    transport = st.text_input("Transport")
    lr_no     = st.text_input("Lr No")
    lr_date   = st.text_input("Lr Date")

st.divider()

num_products = st.selectbox("Number of Products", list(range(1, 11)))
products = []
PRODUCT_OPTIONS = [
    "TRIENZO D","ALL JOINTS","TRIENZO","OMC","CEFSON 200",
    "PENSEN DSR","EPRISPAS","THIOFLEX A","AMIDOL","HYLAGEN G",
    "ACENT P","LIOSON M","ZOLID 600","TRACET",
    "MECSONFORTE","MEGAVIT", "D CART 6", "BRONCHORIL N 600", "SKINTRIX SACHETS", "L GLUTA SACHETS", "CAPIDOL GEL",
    "BRONCHORIL N 100", " RESPIFORTE "
]
for i in range(num_products):
    st.markdown(f"**Product {i+1}**")
    c1 = st.columns(6)
    pname_option = c1[0].selectbox(
        "Product Name",
        options=[""] + PRODUCT_OPTIONS + ["Other"],
        key=f"pn{i}"
    )

    if pname_option == "Other":
        pname = c1[0].text_input("Enter Product Name", key=f"pn_other{i}")
    else:
        pname = pname_option

    pname = (pname or "").upper()
    packing = c1[1].selectbox(
        "Packing",
        options=[
            "1 x 10", "2 x 10", "3 x 10", "4 x 10", "5 x 10",
            "6 x 10", "7 x 10", "8 x 10", "9 x 10", "10 x 10"
        ],
        key=f"pk{i}"
    )
    hsn     = c1[2].text_input("HSN", key=f"hs{i}")
    batch   = c1[3].text_input("Batch", key=f"ba{i}")
    exp_date = c1[4].date_input("EXP", value=None, key=f"ex{i}")
    exp = exp_date.strftime("%m/%y") if exp_date else ""

    qty = c1[5].number_input("QTY", value=None, placeholder="Enter", key=f"qt{i}")

    c2 = st.columns(6)
    free = c2[0].number_input("FREE",  value=None, placeholder="Enter", key=f"fr{i}")
    ptr  = c2[1].number_input("PTR",   value=None, placeholder="Enter", key=f"pr{i}")
    disc = c2[2].number_input("DISC",  value=None, placeholder="Enter", key=f"di{i}")
    mrp  = c2[3].number_input("MRP",   value=None, placeholder="Enter", key=f"mr{i}")
    cgst = 2.5
    sgst = 2.5

    c2[4].text_input("CGST%", value="2.5", disabled=True, key=f"cg{i}")
    c2[5].text_input("SGST%", value="2.5", disabled=True, key=f"sg{i}")
    qty  = qty  or 0
    ptr  = ptr  or 0
    disc = disc or 0
    free = free or 0
    mrp  = mrp  or 0
    taxable  = qty * (ptr - disc)
    cgst_amt = taxable * cgst / 100
    sgst_amt = taxable * sgst / 100

    products.append({
        "Product": pname, "Packing": packing, "HSN": hsn,
        "Batch": batch, "EXP": exp, "Qty": qty, "Free": free,
        "PTR": ptr, "Discount": disc, "MRP": mrp,
        "Taxable": taxable, "CGST%": float(cgst), "SGST%": float(sgst),
        "CGST Amt": cgst_amt, "SGST Amt": sgst_amt,
        "Total": taxable + cgst_amt + sgst_amt,
    })

df = pd.DataFrame(products)

st.divider()

if "excel_file" not in st.session_state:
    st.session_state.excel_file = None
if "pdf_file" not in st.session_state:
    st.session_state.pdf_file = None

valid = any(p["Taxable"] > 0 for p in products)

if st.button("📄 Generate Invoice", disabled=not valid, type="primary"):
    buyer_text = BUYERS[buyer_key]
    st.session_state.excel_file = create_excel(
        df, buyer_key, buyer_text, inv_no, date_val, transport, lr_no, lr_date
    )
    st.session_state.pdf_file = create_pdf(
        df, buyer_text, inv_no, date_val, transport, lr_no, lr_date
    )

if st.session_state.excel_file and st.session_state.pdf_file:
    st.success("✅ Invoice generated! Download below:")
    fname = re.sub(r"[^a-zA-Z0-9]", "_", name or inv_no or "invoice")
    dl1, dl2 = st.columns(2)
    with dl1:
        st.download_button(
            "⬇️ Download Excel",
            st.session_state.excel_file,
            file_name=f"{fname}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with dl2:
        st.download_button(
            "⬇️ Download PDF",
            st.session_state.pdf_file,
            file_name=f"{fname}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
